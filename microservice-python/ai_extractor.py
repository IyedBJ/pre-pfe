import sys
import os
import shutil
import json
import re

from pdf_utils import extract_text_pdf, extract_text_ocr, is_pdf_scanned
from nlp_utils import clean_text
from ollama_utils import extract_ollama

def clean_number(value):
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = re.sub(r'[^\d,\.\s]', '', value).strip()
        if not s: return None
        s = s.replace(' ', '')
        
        if ',' in s and '.' in s:
            if s.rfind(',') > s.rfind('.'):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            s = s.replace(',', '.')
        elif '.' in s:
            if len(s.split('.')[-1]) == 3 and len(s) > 4:
                s = s.replace('.', '')
                
        try:
            return float(s)
        except:
            return None
    return None

def extract_excel_data(file_path, ext=None):
    if not ext:
        ext = os.path.splitext(file_path)[1].lower()
    
    all_numbers = []
    temp_path = None
    
    try:
        if not file_path.lower().endswith(tuple(['.xlsx', '.xls', '.xlsm', '.xltx', '.xltm'])):
            temp_path = file_path + ext
            shutil.copy(file_path, temp_path)
            target_path = temp_path
        else:
            target_path = file_path

        result_data = {"total": 0}

        if ext == '.xlsx':
            try:
                import openpyxl
            except ImportError:
                return {"error": "Module 'openpyxl' non trouvé."}
            
            wb = openpyxl.load_workbook(target_path, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for col_idx, value in enumerate(row):
                        num = clean_number(value)
                        if num is not None: all_numbers.append(num)
                        
                        if isinstance(value, str) and value.strip().lower() == "date":
                            for offset in range(1, 4):
                                if col_idx + offset < len(row) and row[col_idx + offset]:
                                    val_date = row[col_idx + offset]
                                    import datetime
                                    if isinstance(val_date, (datetime.datetime, datetime.date)):
                                        result_data["date"] = f"{val_date.year}-{val_date.month:02d}"
                                    elif isinstance(val_date, str):
                                        import re
                                        date_pattern = r'(\d{1,2})[/\-_](\d{1,2})[/\-_](\d{4})'
                                        m = re.search(date_pattern, val_date)
                                        if m:
                                            m1, m2, yr = int(m.group(1)), int(m.group(2)), m.group(3)
                                            mo = m1 if m2 > 12 else m2
                                            result_data["date"] = f"{yr}-{mo:02d}"

                        if isinstance(value, str) and "total à verser" in value.lower():
                            for offset in range(1, 4):
                                if col_idx + offset < len(row):
                                    n = clean_number(row[col_idx + offset])
                                    if n: return {"total": n}

                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    for col_idx, value in enumerate(row, 1):
                        if isinstance(value, str) and "total en euro" in value.lower():
                            last_found = 0
                            for r in range(row_idx + 1, sheet.max_row + 1):
                                val = sheet.cell(row=r, column=col_idx).value
                                n = clean_number(val)
                                if n: last_found = n
                            if last_found: result_data["total"] = last_found

        elif ext == '.xls':
            try:
                import xlrd
            except ImportError:
                return {"error": "Module 'xlrd' non trouvé."}
                
            wb = xlrd.open_workbook(target_path)
            for sheet in wb.sheets():
                for rx in range(sheet.nrows):
                    row = [sheet.cell_value(rx, cx) for cx in range(sheet.ncols)]
                    for cx, value in enumerate(row):
                        num = clean_number(value)
                        if num is not None: all_numbers.append(num)
                        
                        if isinstance(value, str) and value.strip().lower() == "date":
                            for offset in range(1, 4):
                                if cx + offset < len(row) and row[cx + offset]:
                                    val_date = row[cx + offset]
                                    try:
                                        if isinstance(val_date, float):
                                            dt = xlrd.xldate_as_datetime(val_date, wb.datemode)
                                            result_data["date"] = f"{dt.year}-{dt.month:02d}"
                                        elif isinstance(val_date, str):
                                            import re
                                            date_pattern = r'(\d{1,2})[/\-_](\d{1,2})[/\-_](\d{4})'
                                            m = re.search(date_pattern, val_date)
                                            if m:
                                                m1, m2, yr = int(m.group(1)), int(m.group(2)), m.group(3)
                                                mo = m1 if m2 > 12 else m2
                                                result_data["date"] = f"{yr}-{mo:02d}"
                                    except:
                                        pass

                        if isinstance(value, str) and "total à verser" in value.lower():
                            for offset in range(1, 4):
                                if cx + offset < len(row):
                                    n = clean_number(row[cx + offset])
                                    if n: return {"total": n}

                for rx in range(sheet.nrows):
                    for cx in range(sheet.ncols):
                        value = sheet.cell_value(rx, cx)
                        if isinstance(value, str) and "total en euro" in value.lower():
                            last_found = 0
                            for r in range(rx + 1, sheet.nrows):
                                n = clean_number(sheet.cell_value(r, cx))
                                if n: last_found = n
                            if last_found: result_data["total"] = last_found
        
        if result_data.get("total") == 0 and all_numbers:
            result_data["total"] = float(max(all_numbers))
            
        return result_data
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": f"Erreur Excel: {str(e)}"}
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)

def extract_financial_data(file_path, file_ext):
    if file_ext in ['.xlsx', '.xls']:
        return extract_excel_data(file_path, file_ext)
    
    if is_pdf_scanned(file_path):
        text = extract_text_ocr(file_path) 
    else:
        text = extract_text_pdf(file_path)

    text = clean_text(text)
    data = extract_ollama(text)
    return data

if __name__ == "__main__":
    import os
    file_path = sys.argv[1]
    file_ext = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(file_path)[1].lower()
    
    
    result = extract_financial_data(file_path, file_ext)
    
    print(json.dumps(result))             
 