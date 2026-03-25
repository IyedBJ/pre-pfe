import sys
import os
import zipfile
import shutil
import json
import tempfile
import re

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from ai_extractor import extract_financial_data, extract_excel_data
    from pdf_utils import extract_text_pdf, is_pdf_scanned, extract_text_ocr
except ImportError as e:
    print(json.dumps({"error": f"Import error: {str(e)}"}))
    sys.exit(1)


FRENCH_MONTHS = {
    'janvier': '01', 'fevrier': '02', 'février': '02', 'mars': '03',
    'avril': '04', 'mai': '05', 'juin': '06', 'juillet': '07',
    'aout': '08', 'août': '08', 'septembre': '09', 'octobre': '10',
    'novembre': '11', 'decembre': '12', 'décembre': '12'
}

def find_date_in_text(text):

    if not text:
        return None
    
    text_str = str(text)
    
    #"Septembre 2020", "Période : Septembre 2020")
    fr_date_pattern = r'([a-zA-Zéû]+)[^\w\d]*(\d{4})'
    for match in re.finditer(fr_date_pattern, text_str.lower()):
        month_str, year = match.groups()
        if month_str in FRENCH_MONTHS and 1990 <= int(year) <= 2100:
            return f"{year}-{FRENCH_MONTHS[month_str]}"

    # YYYY-MM-DD
    iso_match = re.search(r'(\d{4})[-/](\d{2})[-/]\d{2}', text_str)
    if iso_match:
        year, month = iso_match.group(1), iso_match.group(2)
        if 1 <= int(month) <= 12 and 1990 <= int(year) <= 2100:
            return f"{year}-{month}"
    
    # MM/DD/YYYY or DD/MM/YYYY
    date_pattern = r'(\d{1,2})[/\-_](\d{1,2})[/\-_](\d{4})'
    matches = re.findall(date_pattern, text_str)
    for m in matches:
        m1, m2, year = int(m[0]), int(m[1]), m[2]
        if 1990 <= int(year) <= 2100:
            if m2 > 12:
                month = m1
            else:
                month = m2 
            if 1 <= month <= 12:
                return f"{year}-{month:02d}"
    return None


def find_date_in_excel_cells(file_path, ext):

    import datetime
    try:
        if ext == '.xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is None:
                            continue
                        if isinstance(cell, (datetime.datetime, datetime.date)):
                            return f"{cell.year}-{cell.month:02d}"
                        if isinstance(cell, str):
                            result = find_date_in_text(cell)
                            if result:
                                return result
        elif ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(file_path)
            for sheet in wb.sheets():
                for rx in range(sheet.nrows):
                    for cx in range(sheet.ncols):
                        cell = sheet.cell(rx, cx)
                        if cell.ctype == 3:
                            try:
                                dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                                return f"{dt.year}-{dt.month:02d}"
                            except:
                                pass
                        elif cell.ctype == 1:  # text
                            result = find_date_in_text(cell.value)
                            if result:
                                return result
    except Exception as e:
        print(f"[Python] Excel date scan error: {e}", file=sys.stderr)
    return None



def process_zip(zip_path, employee_name, file_type):

    results = []
    temp_dir = tempfile.mkdtemp()

    print(f"[Python] Extracting ZIP: {zip_path}", file=sys.stderr)
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            print(f"[Python] ZIP contains {len(zf.namelist())} files: {zf.namelist()}", file=sys.stderr)
            zf.extractall(temp_dir)
    except zipfile.BadZipFile as e:
        print(f"[Python] ERROR - Bad ZIP file: {e}", file=sys.stderr)
        return []
    except Exception as e:
        print(f"[Python] ERROR - Cannot open ZIP: {e}", file=sys.stderr)
        return []

    for root, dirs, files in os.walk(temp_dir):
        for filename in files:
            file_path = os.path.join(root, filename)
            ext = os.path.splitext(filename)[1].lower()

            if ext not in ['.pdf', '.xlsx', '.xls']:
                print(f"[Python] Skipping non-supported file: {filename}", file=sys.stderr)
                continue

            print(f"[Python] Processing file: {filename} (ext={ext})", file=sys.stderr)

            try:
                data = extract_financial_data(file_path, ext)
                if "error" in data:
                    print(f"[Python] Extraction error for {filename}: {data['error']}", file=sys.stderr)
                    continue

                data["filename"] = filename
                data["file_type"] = file_type

                detected_month = data.get("date") or find_date_in_text(filename)
                
                if not detected_month and ext in ['.xlsx', '.xls']:
                    detected_month = find_date_in_excel_cells(file_path, ext)
                    if detected_month:
                        print(f"[Python] Date found in Excel cells: {detected_month}", file=sys.stderr)
                elif not detected_month and ext == '.pdf':
                    try:
                        if is_pdf_scanned(file_path):
                            text = extract_text_ocr(file_path)
                        else:
                            text = extract_text_pdf(file_path)
                        detected_month = find_date_in_text(text)
                    except:
                        pass

                data["date_group"] = detected_month or "Unknown"
                print(f"[Python] SUCCESS {filename}: total={data.get('total', data.get('net_paye'))}, month={data['date_group']}", file=sys.stderr)
                results.append(data)

            except Exception as e:
                print(f"[Python] Error extracting data from {filename}: {e}", file=sys.stderr)
                import traceback
                traceback.print_exc(file=sys.stderr)

    shutil.rmtree(temp_dir, ignore_errors=True)
    print(f"[Python] Finished. Found {len(results)} items.", file=sys.stderr)
    
    return results


if __name__ == "__main__":
    try:
        if len(sys.argv) < 3:
            print(json.dumps({"error": "Usage: zip_processor.py <zip_path> <employee_name> [file_type]"}))
            sys.exit(1)

        zip_path = sys.argv[1]
        employee_name = sys.argv[2]
        file_type = sys.argv[3] if len(sys.argv) > 3 else "unknown"

        print(f"[Python] Start: employee='{employee_name}', type='{file_type}', file='{zip_path}'", file=sys.stderr)

        if not os.path.exists(zip_path):
            print(json.dumps({"error": f"File not found: {zip_path}"}))
            sys.exit(1)

        results = process_zip(zip_path, employee_name, file_type)
        print(json.dumps(results))

    except Exception as e:
        import traceback
        traceback.print_exc(file=sys.stderr)
        print(json.dumps({"error": str(e)}))
        sys.exit(1)
